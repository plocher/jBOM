"""
Inventory file loader for jBOM.

Handles loading inventory data from multiple file formats:
- CSV (.csv)
- Excel (.xlsx, .xls)
- Apple Numbers (.numbers)
"""

import csv
import logging
import warnings
from pathlib import Path
from typing import List, Dict, Optional, Union

# warnings imported at line 12 already
from jbom.common.component_classification import normalize_component_type
from jbom.common.component_id import is_current_version, make_component_id
from jbom.common.types import InventoryItem, DEFAULT_PRIORITY
from jbom.common.value_parsing import (
    TYPED_PARAMETRIC_COLUMNS_BY_CATEGORY,
    UNCLASSIFIED_CATEGORIES,
    decode_typed_parametric,
)
from jbom.common.synonym_normalization import first_non_empty_alias_value
from jbom.config.defaults import DefaultsConfig, get_defaults
from jbom.services.jlc_loader import JLCPrivateInventoryLoader

log = logging.getLogger(__name__)

_ROW_TYPE_ITEM = "ITEM"
_ROW_TYPE_COMPONENT = "COMPONENT"


# Suppress specific Numbers version warning
warnings.filterwarnings(
    "ignore", message="Numbers version 14.3 not tested with this version"
)

# Optional imports for spreadsheet support
try:
    import openpyxl

    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

try:
    from numbers_parser import Document as NumbersDocument

    NUMBERS_SUPPORT = True
except ImportError:
    NUMBERS_SUPPORT = False


def _resolve_component_id(
    row_type: str,
    stored_id: str,
    row: Dict[str, str],
) -> str:
    """Return a current-version ComponentID for a COMPONENT row.

    If *stored_id* is already at the current encoding version it is returned
    unchanged.  Otherwise the ID is regenerated from the row's field data so
    that stale IDs (e.g. legacy ``REQ1|...`` format, wrong version number, or
    blank) are transparently upgraded on load without requiring manual action.

    ITEM rows are returned as-is (they do not carry a ComponentID).
    """
    if row_type != _ROW_TYPE_COMPONENT:
        return stored_id
    if is_current_version(stored_id):
        return stored_id
    return make_component_id(
        category=row.get("Category", ""),
        value=row.get("Value", ""),
        package=row.get("Package", ""),
        tolerance=row.get("Tolerance", ""),
        voltage=row.get("Voltage") or row.get("V", ""),
        amperage=row.get("Current") or row.get("A", ""),
        wattage=row.get("Power") or row.get("W", ""),
        component_type=row.get("Type", ""),
    )


class InventoryReader:
    """Loads inventory data from various file formats."""

    def __init__(self, inventory_paths: Union[Path, List[Path]]):
        """Initialize loader with path(s) to inventory file(s).

        Args:
            inventory_paths: Path or list of Paths to inventory file(s)
        """
        if isinstance(inventory_paths, Path):
            self.inventory_paths = [inventory_paths]
        else:
            self.inventory_paths = inventory_paths

        self.inventory: List[InventoryItem] = []
        self.inventory_fields: List[str] = []
        self._defaults_profile: DefaultsConfig | None = None

    def _get_defaults_profile(self) -> DefaultsConfig:
        """Return the active defaults profile for this reader instance."""

        if self._defaults_profile is None:
            self._defaults_profile = get_defaults()
        return self._defaults_profile

    def load(self) -> tuple[List[InventoryItem], List[str]]:
        """Load inventory from all provided files.

        Returns:
            Tuple of (aggregated inventory items list, aggregated field names list)
        """
        for path in self.inventory_paths:
            self._load_file(path)

        return self.inventory, list(set(self.inventory_fields))

    def _load_file(self, path: Path):
        """Load a single inventory file."""
        file_extension = path.suffix.lower()

        if file_extension == ".csv":
            self._load_csv_inventory(path)
        elif file_extension in [".xlsx", ".xls"]:
            if not EXCEL_SUPPORT:
                raise ImportError(
                    "Excel support requires openpyxl package. Install with: pip install openpyxl"
                )
            # Try JLC Loader first for Excel files
            if self._try_load_jlc(path):
                return

            self._load_excel_inventory(path)
        elif file_extension == ".numbers":
            if not NUMBERS_SUPPORT:
                raise ImportError(
                    "Numbers support requires numbers-parser package. Install with: pip install numbers-parser"
                )
            self._load_numbers_inventory(path)
        else:
            raise ValueError(
                f"Unsupported inventory file format: {file_extension}. Supported formats: .csv, .xlsx, .xls, .numbers"
            )

    def _try_load_jlc(self, path: Path) -> bool:
        """Attempt to load as JLC Private Inventory. Returns True if successful."""
        try:
            loader = JLCPrivateInventoryLoader(path)
            # We need to peek at headers or just try loading
            # The loader handles validation internally
            items, fields = loader.load()

            # Merge results
            self.inventory.extend(items)
            self.inventory_fields.extend(fields)
            return True
        except (ValueError, KeyError):
            # Not a JLC file, or missing headers
            return False

    def _load_csv_inventory(self, path: Path):
        """Load inventory from CSV file"""
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            headers = reader.fieldnames or []
            self._process_inventory_data(headers, rows, source="CSV", source_file=path)

    def _load_excel_inventory(self, path: Path):
        """Load inventory from Excel file (.xlsx or .xls)"""
        workbook = openpyxl.load_workbook(path, data_only=True)
        # Use the first worksheet
        worksheet = workbook.active

        # Find the header row by looking for a characteristic schema column
        header_row = None
        start_col = 1

        for row_num in range(1, min(10, worksheet.max_row + 1)):  # Check first 10 rows
            for col_num in range(
                1, min(10, worksheet.max_column + 1)
            ):  # Check first 10 columns
                cell_value = worksheet.cell(row_num, col_num).value
                if cell_value and str(cell_value).strip().upper() in {
                    "IPN",
                    "CATEGORY",
                    "ROWTYPE",
                }:
                    header_row = row_num
                    start_col = col_num
                    break
            if header_row:
                break

        if not header_row:
            raise ValueError(
                f"Could not find an inventory header row in Excel file {path}. "
                "Expected columns such as RowType, Category, or IPN."
            )

        # Get headers from the identified header row
        headers = []
        col_num = start_col
        while col_num <= worksheet.max_column:
            cell_value = worksheet.cell(header_row, col_num).value
            if cell_value is not None and str(cell_value).strip():
                headers.append(str(cell_value).strip())
                col_num += 1
            else:
                # Stop when we hit an empty header cell
                break

        # Get data rows
        rows = []
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            row_data = {}
            has_data = False

            for col_offset, header in enumerate(headers):
                col_num = start_col + col_offset
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row_num, col_num).value
                    value_str = (
                        str(cell_value).strip() if cell_value is not None else ""
                    )
                    row_data[header] = value_str
                    if value_str:  # Check if this row has any data
                        has_data = True
                else:
                    row_data[header] = ""

            # Only add rows that have some data
            if has_data:
                rows.append(row_data)

        workbook.close()
        self._process_inventory_data(headers, rows, source="Excel", source_file=path)

    def _load_numbers_inventory(self, path: Path):
        """Load inventory from Apple Numbers file"""
        doc = NumbersDocument(path)
        # Get the first table from the first sheet
        if not doc.sheets:
            raise ValueError("No sheets found in Numbers file")

        sheet = doc.sheets[0]
        if not sheet.tables:
            raise ValueError("No tables found in first sheet of Numbers file")

        table = sheet.tables[0]

        # Find the header row by looking for a characteristic schema column
        header_row_idx = None
        start_col = None

        for row_idx in range(min(10, table.num_rows)):  # Check first 10 rows
            for col_idx in range(min(10, table.num_cols)):  # Check first 10 columns
                cell = table.cell(row_idx, col_idx)
                if cell.value and str(cell.value).strip().upper() in {
                    "IPN",
                    "CATEGORY",
                    "ROWTYPE",
                }:
                    header_row_idx = row_idx
                    start_col = col_idx
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            raise ValueError(
                "Could not find an inventory header row in Numbers file. "
                "Expected columns such as RowType, Category, or IPN."
            )

        # Get headers from the identified header row
        headers = []

        for col_idx in range(start_col, table.num_cols):
            cell = table.cell(header_row_idx, col_idx)
            if cell.value is not None and str(cell.value).strip():
                headers.append(str(cell.value).strip())
            else:
                # Stop when we hit an empty header cell
                break

        # Get data rows
        rows = []
        for row_idx in range(header_row_idx + 1, table.num_rows):
            row_data = {}
            has_data = False

            for col_offset, header in enumerate(headers):
                col_idx = start_col + col_offset
                if col_idx < table.num_cols:
                    cell = table.cell(row_idx, col_idx)
                    value_str = (
                        str(cell.value).strip() if cell.value is not None else ""
                    )
                    row_data[header] = value_str
                    if value_str:  # Check if this row has any data
                        has_data = True
                else:
                    row_data[header] = ""

            # Only add rows that have some data
            if has_data:
                rows.append(row_data)

        self._process_inventory_data(headers, rows, source="Numbers", source_file=path)

    def _process_inventory_data(
        self,
        headers: List[str],
        rows: List[Dict[str, str]],
        source: str = "Unknown",
        source_file: Optional[Path] = None,
    ):
        """Process inventory data from any source format into InventoryItem objects"""
        # Validate required headers
        required_headers = ["Category"]
        header_upper = [h.upper() for h in headers]
        missing_headers = [
            req for req in required_headers if req.upper() not in header_upper
        ]

        if missing_headers:
            raise ValueError(
                f"Inventory file is missing required columns: {', '.join(missing_headers)}. "
                f"Found columns: {', '.join(headers)}"
            )

        # Clean up field names - replace newlines and extra whitespace
        self.inventory_fields = []
        for field in headers:
            if field:
                # Replace newlines with spaces and normalize whitespace
                clean_field = " ".join(
                    field.replace("\n", " ").replace("\r", " ").split()
                )
                self.inventory_fields.append(clean_field)

        for row in rows:
            row_type = str(row.get("RowType", _ROW_TYPE_ITEM)).strip().upper()
            if row_type not in {_ROW_TYPE_ITEM, _ROW_TYPE_COMPONENT}:
                row_type = _ROW_TYPE_ITEM

            ipn = str(row.get("IPN", "")).strip()
            if row_type == _ROW_TYPE_ITEM and not ipn:
                continue

            # No need to parse stocking info - Priority column handles all ranking
            category = row.get("Category", "")
            value = row.get("Value", "")
            row_context = (
                ipn
                or str(row.get("ComponentID", "")).strip()
                or str(row.get("UUID", "")).strip()
                or str(value).strip()
                or "<row>"
            )
            (
                effective_category,
                decode_category,
            ) = self._resolve_category_for_typed_decode(
                source_category=category,
                row=row,
                context=row_context,
            )
            item = InventoryItem(
                ipn=ipn,
                keywords=row.get("Keywords", ""),
                category=effective_category,
                description=row.get("Description", ""),
                smd=row.get("SMD", ""),
                value=value,
                type=row.get("Type", ""),
                tolerance=row.get("Tolerance", ""),
                voltage=self._get_canonical_electrical_value(
                    row,
                    canonical="voltage",
                ),
                amperage=self._get_canonical_electrical_value(
                    row,
                    canonical="current",
                ),
                wattage=self._get_canonical_electrical_value(
                    row,
                    canonical="power",
                ),
                row_type=row_type,
                component_id=_resolve_component_id(
                    row_type, str(row.get("ComponentID", "")).strip(), row
                ),
                supplier=str(row.get("Supplier", "")).strip(),
                spn=str(row.get("SPN", "")).strip(),
                manufacturer=self._get_canonical_profile_value(
                    row,
                    canonical="manufacturer",
                    fallback_keys=["Manufacturer"],
                ),
                mfgpn=self._get_canonical_profile_value(
                    row,
                    canonical="mpn",
                    fallback_keys=[
                        "MFGPN",
                        "MPN",
                        "Manufacturer Part Number",
                    ],
                ),
                datasheet=row.get("Datasheet", ""),
                package=row.get("Package", ""),
                uuid=row.get("UUID", ""),
                priority=self._parse_priority(
                    row.get("Priority", str(DEFAULT_PRIORITY))
                ),
                # Typed parametric fields decoded at intake (#90).
                # Each call passes the TARGET field's category so the helper knows
                # which column to look in and which parse function to apply.
                resistance=(
                    decode_typed_parametric("RES", value, row)
                    if decode_category == "RES"
                    else None
                ),
                capacitance=(
                    decode_typed_parametric("CAP", value, row)
                    if decode_category == "CAP"
                    else None
                ),
                inductance=(
                    decode_typed_parametric("IND", value, row)
                    if decode_category == "IND"
                    else None
                ),
                name=(row.get("Name") or "").strip(),
                aliases=first_non_empty_alias_value(row, ["Aliases", "Alias"]),
                dnp=self._is_truthy_inventory_flag(
                    first_non_empty_alias_value(
                        row,
                        ["DNP", "Do Not Populate", "DoNotPopulate"],
                    )
                ),
                # KiCad harvest fidelity fields: populated when columns are present;
                # absent columns leave fields as empty string (no compat shims).
                footprint_full=row.get("footprint_full", ""),
                symbol_lib=row.get("symbol_lib", ""),
                symbol_name=row.get("symbol_name", ""),
                pins=row.get("Pins", ""),
                pitch=row.get("Pitch", ""),
                source=source,
                source_file=source_file,
                raw_data=row,
            )
            self.inventory.append(item)

    def _resolve_category_for_typed_decode(
        self,
        *,
        source_category: str,
        row: Dict[str, str],
        context: str,
    ) -> tuple[str, Optional[str]]:
        """Return (effective_category, decode_category) for typed decode gating."""

        normalized_category = normalize_component_type(source_category)
        if normalized_category in TYPED_PARAMETRIC_COLUMNS_BY_CATEGORY:
            return source_category, normalized_category
        if normalized_category not in UNCLASSIFIED_CATEGORIES:
            return source_category, None

        populated_categories = [
            category
            for category, column in TYPED_PARAMETRIC_COLUMNS_BY_CATEGORY.items()
            if str(row.get(column, "")).strip()
        ]

        if len(populated_categories) == 1:
            promoted_category = populated_categories[0]
            return promoted_category, promoted_category

        if len(populated_categories) > 1:
            populated_columns = ", ".join(
                TYPED_PARAMETRIC_COLUMNS_BY_CATEGORY[category]
                for category in populated_categories
            )
            log.warning(
                "Ambiguous typed parametric promotion for unclassified inventory row '%s': "
                "multiple typed attributes set (%s). Skipping typed decode.",
                context,
                populated_columns,
            )

        return source_category, None

    def _get_canonical_electrical_value(
        self, row: Dict[str, str], *, canonical: str
    ) -> str:
        """Resolve an electrical attribute via defaults-profile synonym mappings.

        No hardcoded aliases are applied here by design: profile configuration
        is authoritative, and intentional omission means aliases are disabled.
        """

        return self._get_canonical_profile_value(row, canonical=canonical)

    def _get_canonical_profile_value(
        self,
        row: Dict[str, str],
        *,
        canonical: str,
        fallback_keys: Optional[List[str]] = None,
    ) -> str:
        """Resolve a canonical value via defaults-profile synonym mappings."""

        keys: List[str] = []
        config = self._get_defaults_profile().get_field_synonym_config(canonical)
        if config is not None:
            keys.extend([config.display_name, *config.synonyms])
        keys.extend(fallback_keys or [])

        deduped_keys: List[str] = []
        seen: set[str] = set()
        for key in keys:
            normalized = key.strip()
            if not normalized or normalized.lower() in seen:
                continue
            seen.add(normalized.lower())
            deduped_keys.append(normalized)
        return first_non_empty_alias_value(row, deduped_keys)

    def _get_first_value(self, row: Dict[str, str], keys: List[str]) -> str:
        """Get the first non-empty value from row matching any of the keys"""
        for key in keys:
            if val := row.get(key):
                return val
        return ""

    def _parse_priority(self, priority_str: str) -> int:
        """Parse priority value from CSV, defaulting to DEFAULT_PRIORITY if invalid"""
        try:
            return (
                int(priority_str.strip()) if priority_str.strip() else DEFAULT_PRIORITY
            )
        except (ValueError, AttributeError):
            return DEFAULT_PRIORITY

    @staticmethod
    def _is_truthy_inventory_flag(value: str) -> bool:
        """Return True when a spreadsheet flag value should be interpreted as enabled."""
        normalized = str(value or "").strip().lower()
        if not normalized:
            return False
        return normalized in {
            "1",
            "true",
            "t",
            "yes",
            "y",
            "x",
            "dnp",
            "do not populate",
        }
