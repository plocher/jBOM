#!/usr/bin/env python3
"""
Automated Inventory Value Normalization Script

Applies the Unicode normalization fixes identified in the analysis to the source inventory file.
Supports CSV, Excel, and Numbers formats.

Usage:
    python apply_inventory_fixes.py examples/example-INVENTORY.numbers --output examples/example-INVENTORY-fixed.numbers --dry-run
    python apply_inventory_fixes.py examples/example-INVENTORY.csv --output examples/example-INVENTORY-fixed.csv --apply
"""

import sys
import os
import csv
import argparse
from pathlib import Path
from typing import Dict, Any, Optional

# Add jBOM to path
sys.path.insert(0, "/Users/jplocher/Dropbox/KiCad/jBOM/src")

from jbom.loaders.inventory import InventoryLoader  # noqa: E402


class InventoryFixer:
    """Applies Unicode normalization fixes to inventory files."""

    # Define the normalization mapping from analysis
    VALUE_FIXES = {
        # Resistor values - remove Ω symbol
        "200Ω": "200",
        "220Ω": "220",
        "330Ω": "330",
        "470Ω": "470",
        "1kΩ": "1k",
        "2.2kΩ": "2.2k",
        "3.3kΩ": "3.3k",
        "4.7kΩ": "4.7k",
        "5.6kΩ": "5.6k",
        "10kΩ": "10k",
        "22kΩ": "22k",
        "100kΩ": "100k",
        "1MΩ": "1M",
        "2.2MΩ": "2.2M",
        # Capacitor values - normalize µF to uF
        "0.01µF": "0.01uF",
        "0.1µF": "0.1uF",
        "1µF": "1uF",
        "10µF": "10uF",
        "100µF": "100uF",
        "220µF": "220uF",
        # Also handle alternative unicode characters
        "0.01μF": "0.01uF",
        "0.1μF": "0.1uF",
        "1μF": "1uF",
        "10μF": "10uF",
        "100μF": "100uF",
        "220μF": "220uF",
    }

    def __init__(self, input_file: Path, output_file: Optional[Path] = None):
        """Initialize the inventory fixer.

        Args:
            input_file: Path to input inventory file
            output_file: Path to output file (if None, modifies input file in-place with backup)
        """
        self.input_file = input_file
        self.output_file = output_file
        self.in_place_mode = output_file is None

    def analyze_fixes_needed(self) -> Dict[str, Any]:
        """Analyze the inventory and identify what fixes are needed.

        Returns:
            Dictionary with analysis results
        """
        print(f"Analyzing inventory: {self.input_file}")

        # Load inventory
        loader = InventoryLoader(self.input_file)
        items, fields = loader.load()

        # Analyze values that need fixing
        fixes_needed = []
        categories_affected = set()

        for item in items:
            if item.value and item.value in self.VALUE_FIXES:
                fixes_needed.append(
                    {
                        "ipn": item.ipn,
                        "category": item.category,
                        "old_value": item.value,
                        "new_value": self.VALUE_FIXES[item.value],
                        "description": item.description,
                    }
                )
                categories_affected.add(item.category)

        return {
            "total_items": len(items),
            "fixes_needed": fixes_needed,
            "fixes_count": len(fixes_needed),
            "categories_affected": sorted(categories_affected),
            "fields": fields,
        }

    def apply_fixes_csv(self, dry_run: bool = True) -> bool:
        """Apply fixes to CSV file with backup support.

        Args:
            dry_run: If True, don't actually write changes

        Returns:
            True if successful
        """
        if self.in_place_mode:
            print(f"Modifying CSV file in-place: {self.input_file}")
        else:
            print(f"Loading CSV file: {self.input_file}")
            print(f"Will save modified version to: {self.output_file}")
        print(
            "Note: CSV format cannot preserve formulas (use Excel/Numbers for formulas)"
        )

        fixes_applied = 0

        # Read CSV file with error handling
        try:
            with open(self.input_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                fieldnames = reader.fieldnames
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return False

        if not fieldnames or "Value" not in fieldnames:
            print("Error: CSV file must have a 'Value' column")
            if fieldnames:
                print(f"Available columns: {', '.join(fieldnames)}")
            return False

        # Apply fixes
        for row in rows:
            if row.get("Value") and row["Value"] in self.VALUE_FIXES:
                old_value = row["Value"]
                new_value = self.VALUE_FIXES[old_value]

                ipn = row.get("IPN", "UNKNOWN")
                print(f"  {ipn}: '{old_value}' → '{new_value}'")

                if not dry_run:
                    row["Value"] = new_value
                fixes_applied += 1

        print(f"Found {fixes_applied} values to fix")

        # Write fixed file if not dry run
        if not dry_run and fixes_applied > 0:
            try:
                output_path = self.output_file if self.output_file else self.input_file

                # Create timestamped backup if modifying in place
                if self.in_place_mode:
                    from datetime import datetime

                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    backup_file = (
                        self.input_file.parent
                        / f"{self.input_file.stem}-backup-{timestamp}{self.input_file.suffix}"
                    )
                    print(f"💾 Creating timestamped backup: {backup_file.name}")
                    import shutil

                    shutil.copy2(self.input_file, backup_file)

                with open(output_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)

                print(
                    f"✅ CSV file saved with {fixes_applied} fixes applied to: {output_path}"
                )

                if self.in_place_mode:
                    print(f"📁 Backup saved as: {backup_file.name}")

            except Exception as e:
                print(f"Error saving CSV file: {e}")
                return False

        return True

    def apply_fixes_excel(self, dry_run: bool = True) -> bool:
        """Apply fixes to Excel file preserving formulas and formatting.

        Args:
            dry_run: If True, don't actually write changes

        Returns:
            True if successful
        """
        try:
            import openpyxl
        except ImportError:
            print("Error: openpyxl not available for Excel support")
            print("Install with: pip install openpyxl")
            return False

        if self.in_place_mode:
            print(f"Modifying Excel file in-place: {self.input_file}")
        else:
            print(f"Loading Excel file: {self.input_file}")
            print(f"Will save modified version to: {self.output_file}")
        print("Note: This preserves formulas, formatting, and worksheet structure")

        # Load workbook with data_only=False to preserve formulas
        try:
            workbook = openpyxl.load_workbook(self.input_file, data_only=False)
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False

        worksheet = workbook.active
        if not worksheet:
            print("Error: No active worksheet found in Excel file")
            return False

        # Find Value and IPN columns by scanning header row
        value_col = None
        ipn_col = None

        # Check multiple possible header rows (sometimes data doesn't start at row 1)
        header_row = 1
        for check_row in range(1, min(5, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(check_row, col).value
                if cell_value:
                    header_text = str(cell_value).strip().upper()
                    if header_text == "VALUE":
                        value_col = col
                        header_row = check_row
                    elif header_text == "IPN":
                        ipn_col = col
                        header_row = check_row

            # If we found Value column, use this as header row
            if value_col:
                break

        if not value_col:
            print("Error: Could not find 'Value' column in Excel file")
            print("Available headers in row 1:")
            for col in range(1, min(10, worksheet.max_column + 1)):
                cell_value = worksheet.cell(1, col).value
                if cell_value:
                    print(f"  Column {col}: '{cell_value}'")
            return False

        if not ipn_col:
            print(
                "Warning: Could not find 'IPN' column - using row numbers for identification"
            )

        print(f"Found Value column at position {value_col}, header row {header_row}")

        # Apply fixes to data rows
        fixes_applied = 0

        for row in range(header_row + 1, worksheet.max_row + 1):
            try:
                # Get the value cell
                value_cell = worksheet.cell(row, value_col)
                current_value = (
                    str(value_cell.value).strip() if value_cell.value else ""
                )

                if current_value in self.VALUE_FIXES:
                    new_value = self.VALUE_FIXES[current_value]

                    # Get IPN for reporting
                    if ipn_col:
                        ipn_cell = worksheet.cell(row, ipn_col)
                        ipn = str(ipn_cell.value) if ipn_cell.value else f"Row-{row}"
                    else:
                        ipn = f"Row-{row}"

                    print(f"  {ipn}: '{current_value}' → '{new_value}'")

                    if not dry_run:
                        # Preserve cell formatting by only changing the value
                        # This maintains number formats, colors, fonts, etc.
                        original_style = value_cell._style
                        value_cell.value = new_value
                        # Restore style if it was changed (shouldn't happen but safety first)
                        if value_cell._style != original_style:
                            value_cell._style = original_style

                    fixes_applied += 1

            except Exception as e:
                print(f"Warning: Error processing row {row}: {e}")
                continue

        print(f"Found {fixes_applied} values to fix")

        # Save the modified workbook if not dry run
        if not dry_run and fixes_applied > 0:
            try:
                output_path = self.output_file if self.output_file else self.input_file

                # Create timestamped backup if modifying in place
                if self.in_place_mode:
                    from datetime import datetime

                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    backup_file = (
                        self.input_file.parent
                        / f"{self.input_file.stem}-backup-{timestamp}{self.input_file.suffix}"
                    )
                    print(f"💾 Creating timestamped backup: {backup_file.name}")
                    import shutil

                    shutil.copy2(self.input_file, backup_file)

                print(f"Saving modified Excel file to: {output_path}")
                workbook.save(output_path)
                print(f"✅ Excel file saved with {fixes_applied} fixes applied")
                print(f"Original formulas, formatting, and structure preserved")

                if self.in_place_mode:
                    print(f"📁 Backup saved as: {backup_file.name}")

            except Exception as e:
                print(f"Error saving Excel file: {e}")
                print(
                    f"This might be due to file permissions or Excel format limitations"
                )
                return False

        workbook.close()
        return True

    def apply_fixes_numbers(self, dry_run: bool = True) -> bool:
        """Handle Numbers files with manual workflow or AppleScript automation.

        Args:
            dry_run: If True, don't actually write changes

        Returns:
            True if successful
        """
        print(f"\nNumbers file processing: {self.input_file}")
        
        if dry_run:
            return self._provide_manual_workflow()
            
        # For --apply mode, offer both manual and automated options
        print(f"\nChoose workflow:")
        print(f"  1. Manual workflow (recommended)")
        print(f"  2. AppleScript automation (experimental)")
        
        try:
            choice = input("Enter choice (1 or 2): ").strip()
            if choice == "2":
                return self._apply_with_applescript()
            else:
                return self._provide_manual_workflow()
        except (KeyboardInterrupt, EOFError):
            print("\nAborted.")
            return False

    def _provide_manual_workflow(self) -> bool:
        """Output markdown-style manual Numbers workflow instructions."""
        excel_temp = f"{self.input_file.stem}.xlsx"
        excel_path = self.input_file.parent / excel_temp
        script_path = os.path.abspath(__file__)
        
        print(f"\nOpen the Numbers spreadsheet:")
        print(f"\n```")
        print(f"open {self.input_file}")
        print(f"```")
        print(f"\nUsing the Numbers app window that opens, choose `File->Export To->Excel...`, `[*] One per sheet`, `Save`,")
        print(f"`{self.input_file.parent}/`, `{excel_temp}` and `Export`")
        print(f"Dismiss the Numbers window and run the conversion program:")
        print(f"\n```")
        print(f"python '{script_path}' '{excel_path}' --apply")
        print(f"```")
        print(f"\nOpen the newly created Excel file in Numbers:")
        print(f"\n```")
        print(f"osascript <<HERE")
        print(f'tell application "Numbers"')
        print(f"    activate")
        print(f'    set excelFilePath to "{excel_path}" ')
        print(f"    open (excelFilePath as POSIX file)")
        print(f"end tell")
        print(f"HERE")
        print(f"```")
        
        if self.in_place_mode:
            print(f"\n... and choose `File->Save`, `{self.input_file.parent}`, `{self.input_file.name}`, `Save` and `Replace`")
            validation_file = self.input_file
        else:
            print(f"\n... and choose `File->Save`, `{self.output_file.parent}`, `{self.output_file.name}`, `Save`")
            validation_file = self.output_file
            
        print(f"\nValidate the updated file with:")
        print(f"\n```")
        print(f"python -m jbom inventory-search {validation_file} --dry-run")
        print(f"```")
        
        return True

    def _apply_with_applescript(self) -> bool:
        """Apply fixes to Numbers file using AppleScript automation."""
        import subprocess
        import tempfile
        from datetime import datetime
        
        excel_temp = f"{self.input_file.stem}.xlsx"
        excel_path = self.input_file.parent / excel_temp
        
        print(f"\nAttempting AppleScript automation...")
        
        # Step 1: Export Numbers to Excel via AppleScript
        export_script = f'''
        tell application "Numbers"
            activate
            set numbersFilePath to "{self.input_file}" as POSIX file
            open numbersFilePath
            delay 2
            
            tell front document
                export to file "{excel_path}" as Microsoft Excel
            end tell
            
            close front document
        end tell
        '''
        
        try:
            print(f"Exporting to Excel: {excel_temp}")
            subprocess.run(["osascript", "-e", export_script], check=True, capture_output=True)
            
            if not excel_path.exists():
                raise Exception(f"Excel export failed - file not created: {excel_path}")
                
        except subprocess.CalledProcessError as e:
            print(f"AppleScript export failed: {e}")
            print(f"Falling back to manual workflow...")
            return self._provide_manual_workflow()
        except Exception as e:
            print(f"Export error: {e}")
            print(f"Falling back to manual workflow...")
            return self._provide_manual_workflow()
            
        # Step 2: Apply fixes to Excel file
        print(f"Applying Unicode fixes to Excel file...")
        excel_fixer = InventoryFixer(excel_path, excel_path)
        if not excel_fixer.apply_fixes_excel(dry_run=False):
            print(f"Failed to apply fixes to Excel file")
            return False
            
        # Step 3: Import Excel back to Numbers via AppleScript
        if self.in_place_mode:
            # Create timestamped backup
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            backup_file = (
                self.input_file.parent
                / f"{self.input_file.stem}-backup-{timestamp}{self.input_file.suffix}"
            )
            import shutil
            shutil.copy2(self.input_file, backup_file)
            print(f"Created backup: {backup_file.name}")
            
            output_path = self.input_file
        else:
            output_path = self.output_file
            
        import_script = f'''
        tell application "Numbers"
            activate
            set excelFilePath to "{excel_path}" as POSIX file
            open excelFilePath
            delay 2
            
            tell front document
                save as "{output_path}"
            end tell
            
            close front document
        end tell
        '''
        
        try:
            print(f"Converting back to Numbers: {output_path.name}")
            subprocess.run(["osascript", "-e", import_script], check=True, capture_output=True)
            
            if not output_path.exists():
                raise Exception(f"Numbers import failed - file not created: {output_path}")
                
        except subprocess.CalledProcessError as e:
            print(f"AppleScript import failed: {e}")
            print(f"Fixed Excel file available at: {excel_path}")
            return False
        except Exception as e:
            print(f"Import error: {e}")
            print(f"Fixed Excel file available at: {excel_path}")
            return False
            
        # Clean up temporary Excel file
        try:
            excel_path.unlink()
        except Exception as e:
            print(f"Warning: Could not remove temporary file {excel_path}: {e}")
            
        print(f"\nAppleScript automation completed successfully!")
        print(f"Fixed Numbers file: {output_path}")
        if self.in_place_mode:
            print(f"Backup saved as: {backup_file.name}")
            
        return True

    def _numbers_csv_export_approach(self, dry_run: bool = True) -> bool:
        """Export Numbers file to CSV with fixes applied."""
        print("\nExporting Numbers file to CSV with fixes applied:")

        try:
            from numbers_parser import Document as NumbersDocument
        except ImportError:
            return False

        # Export to CSV with fixes - use output file path if specified
        if self.output_file.suffix.lower() == ".csv":
            csv_output = self.output_file
        elif self.output_file.suffix.lower() == ".numbers":
            # If output specified as .numbers but we're creating CSV, change extension
            csv_output = self.output_file.parent / f"{self.output_file.stem}.csv"
        else:
            csv_output = self.input_file.parent / f"{self.input_file.stem}-fixed.csv"

        doc = NumbersDocument(self.input_file)
        sheet = doc.sheets[0]
        table = sheet.tables[0]

        # Extract all data including formulas (as values)
        headers = []
        for col in range(table.num_cols):
            cell = table.cell(0, col)
            headers.append(str(cell.value) if cell.value else f"Col{col}")

        # Process rows and apply fixes
        rows = []
        for row in range(1, table.num_rows):
            row_data = {}
            for col in range(table.num_cols):
                cell = table.cell(row, col)
                cell_value = str(cell.value) if cell.value else ""

                # Apply fixes to Value column
                if headers[col] == "Value" and cell_value in self.VALUE_FIXES:
                    if not dry_run:
                        cell_value = self.VALUE_FIXES[cell_value]

                row_data[headers[col]] = cell_value
            rows.append(row_data)

        if not dry_run:
            # Write fixed CSV
            with open(csv_output, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)

            print(f"\n📋 Fixed CSV created: {csv_output}")
            print(f"\n📝 To complete the Numbers update:")
            print(f"1. Open your original Numbers file: {self.input_file}")
            print(f"2. Select the data range (excluding headers)")
            print(f"3. Delete the selected data")
            print(f"4. Import/paste data from: {csv_output}")
            print(f"5. Save the Numbers file")
            print(
                f"\nThis approach preserves your table structure while updating the values."
            )

        return True

    def apply_fixes(self, dry_run: bool = True) -> bool:
        """Apply fixes based on file type.

        Args:
            dry_run: If True, don't actually write changes

        Returns:
            True if successful
        """
        suffix = self.input_file.suffix.lower()

        if suffix == ".csv":
            return self.apply_fixes_csv(dry_run)
        elif suffix in [".xlsx", ".xls"]:
            return self.apply_fixes_excel(dry_run)
        elif suffix == ".numbers":
            return self.apply_fixes_numbers(dry_run)
        else:
            print(f"Error: Unsupported file format: {suffix}")
            return False


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Apply Unicode normalization fixes to inventory files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze what fixes are needed
  python apply_inventory_fixes.py examples/example-INVENTORY.numbers --dry-run

  # Apply fixes to CSV
  python apply_inventory_fixes.py examples/example-INVENTORY.csv --apply

  # Apply fixes with custom output
  python apply_inventory_fixes.py input.numbers --output fixed-inventory.csv --apply
        """,
    )

    parser.add_argument(
        "input_file", help="Input inventory file (CSV, Excel, or Numbers)"
    )

    parser.add_argument(
        "--output", "-o", help="Output file path (default: input_file-fixed.ext)"
    )

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--dry-run",
        action="store_true",
        help="Analyze and show what would be fixed without making changes",
    )

    group.add_argument(
        "--apply",
        action="store_true",
        help="Actually apply the fixes to create fixed file",
    )

    args = parser.parse_args()

    # Validate input file
    input_file = Path(args.input_file)
    if not input_file.exists():
        print(f"Error: Input file not found: {input_file}")
        return 1

    output_file = Path(args.output) if args.output else None

    # Create fixer and run analysis
    fixer = InventoryFixer(input_file, output_file)
    analysis = fixer.analyze_fixes_needed()

    print(f"\n=== INVENTORY FIX ANALYSIS ===")
    print(f"Input file: {input_file}")
    print(f"Total items: {analysis['total_items']}")
    print(f"Items needing fixes: {analysis['fixes_count']}")
    print(f"Categories affected: {', '.join(analysis['categories_affected'])}")

    if analysis["fixes_count"] == 0:
        print("\n✅ No fixes needed - inventory values are already normalized!")
        return 0

    print(f"\n=== FIXES TO BE APPLIED ===")
    for fix in analysis["fixes_needed"]:
        print(
            f"  {fix['ipn']} ({fix['category']}): '{fix['old_value']}' → '{fix['new_value']}'"
        )

    if args.dry_run:
        print(f"\n=== DRY RUN COMPLETE ===")
        print(f"To apply these fixes, run with --apply flag")
        return 0

    # Apply fixes
    print(f"\n=== APPLYING FIXES ===")
    if fixer.in_place_mode:
        print(f"Modifying in-place: {input_file}")
        print(f"(Timestamped backup will be created)")
    else:
        print(f"Output file: {fixer.output_file}")

    success = fixer.apply_fixes(dry_run=False)

    if success:
        print(f"\n✅ Fixes applied successfully!")
        print(
            f"Validate with: python -m jbom inventory-search {fixer.output_file} --dry-run"
        )
    else:
        print(f"\n❌ Error applying fixes")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
